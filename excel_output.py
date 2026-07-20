from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _format_workbook(path: Path) -> None:
    wb=load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
        for cell in ws[1]:
            cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='1F4E78'); cell.alignment=Alignment(horizontal='center')
        for col in range(1,ws.max_column+1):
            values=[str(ws.cell(r,col).value or '') for r in range(1,min(ws.max_row,200)+1)]
            width=min(max(max((len(v) for v in values),default=8)+2,10),45)
            ws.column_dimensions[get_column_letter(col)].width=width
    wb.save(path)


def write_master_workbook(path: Path, **sheets: pd.DataFrame) -> None:
    with pd.ExcelWriter(path,engine='openpyxl') as writer:
        for name,df in sheets.items(): df.to_excel(writer,sheet_name=name[:31],index=False)
    _format_workbook(path)
